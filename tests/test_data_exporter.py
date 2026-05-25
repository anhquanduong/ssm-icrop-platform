import unittest
import pandas as pd
import openpyxl
import io
from app.utils.data_exporter import export_history_to_csv, export_history_to_xlsx

class TestDataExporter(unittest.TestCase):
    """
    Test suite for the CSV and XLSX spreadsheet data export module data_exporter.py.
    """

    def setUp(self):
        # Seed a mock history tracker collection
        self.sim_history = {
            "Gerasdorf Maize": pd.DataFrame({
                "DOY": [120, 121],
                "BIOMASS": [150.0, 320.0],
                "LAI": [0.22, 0.48],
                "F_WATER": [1.0, 0.95],
                "F_NUTR": [1.0, 1.0],
                "Scenario": ["Gerasdorf Maize"] * 2
            }),
            "Iowa High N": pd.DataFrame({
                "DOY": [120, 121],
                "BIOMASS": [180.0, 360.0],
                "LAI": [0.25, 0.55],
                "F_WATER": [1.0, 1.0],
                "F_NUTR": [1.0, 1.0],
                "Scenario": ["Iowa High N"] * 2
            })
        }
        
        self.soil_config = {
            "depth_mm": 1500,
            "initial_water_percent": 30.0,
            "pawc_mm_m": 160.0,
            "som_percent": 3.2
        }

    def test_csv_exporter(self):
        """
        Verify that simulation history is successfully flattened and converted to CSV.
        """
        csv_str = export_history_to_csv(self.sim_history)
        self.assertIsNotNone(csv_str)
        self.assertTrue("Scenario" in csv_str)
        self.assertTrue("Gerasdorf Maize" in csv_str)
        self.assertTrue("Iowa High N" in csv_str)
        
        # Verify vertical concatenation length: 2 runs * 2 rows = 4 rows (+ 1 header row)
        lines = [line for line in csv_str.strip().split("\n") if line]
        self.assertEqual(len(lines), 5)

    def test_xlsx_exporter_sheets_and_formatting(self):
        """
        Verify Excel workbook tab structuring, metadata dashboard, headers,
        alternating fills, and number masking.
        """
        xlsx_bytes = export_history_to_xlsx(
            sim_history=self.sim_history,
            soil_config=self.soil_config,
            latitude=48.2830,
            longitude=16.4670,
            crop_name="Maize Hybrid BOKU"
        )
        
        self.assertIsNotNone(xlsx_bytes)
        self.assertTrue(len(xlsx_bytes) > 0)
        
        # Load workbook in-memory using openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        
        # 1. Verify sheet names
        self.assertListEqual(wb.sheetnames, ["Summary Dashboard", "Simulation Matrix Data"])
        
        # 2. Verify Summary Dashboard values and styling
        ws_summary = wb["Summary Dashboard"]
        self.assertTrue(ws_summary.views.sheetView[0].showGridLines)
        self.assertEqual(ws_summary.cell(row=1, column=1).value, "BOKU SSM-iCrop Simulation Summary Report")
        
        # Check metadata properties are printed correctly
        self.assertEqual(ws_summary.cell(row=4, column=1).value, "Active Crop Profile")
        self.assertEqual(ws_summary.cell(row=4, column=2).value, "Maize Hybrid BOKU")
        self.assertEqual(ws_summary.cell(row=5, column=1).value, "Simulation Latitude")
        self.assertEqual(ws_summary.cell(row=5, column=2).value, 48.2830)
        
        # 3. Verify Data Worksheet styling and frozen panes
        ws_data = wb["Simulation Matrix Data"]
        self.assertTrue(ws_data.views.sheetView[0].showGridLines)
        self.assertEqual(ws_data.freeze_panes, "A2")  # Top row frozen
        
        # Verify data header styling
        header_cell = ws_data.cell(row=1, column=1)
        self.assertEqual(header_cell.value, "DOY")
        # Check dark navy fill color matches Hex 1B365D
        self.assertEqual(header_cell.fill.start_color.rgb, "001B365D")
        self.assertEqual(header_cell.font.color.rgb, "00FFFFFF")
        self.assertTrue(header_cell.font.bold)
        
        # Verify data formatting masks
        # Row 2 is the first actual data row
        doy_cell = ws_data.cell(row=2, column=1)
        biomass_cell = ws_data.cell(row=2, column=2)
        lai_cell = ws_data.cell(row=2, column=3)
        fwater_cell = ws_data.cell(row=2, column=4)
        
        self.assertEqual(doy_cell.value, 120)
        self.assertEqual(doy_cell.number_format, "#,##0")
        
        self.assertEqual(biomass_cell.value, 150.0)
        self.assertEqual(biomass_cell.number_format, "#,##0")
        
        self.assertEqual(lai_cell.value, 0.22)
        self.assertEqual(lai_cell.number_format, "0.00")
        
        self.assertEqual(fwater_cell.value, 1.0)
        self.assertEqual(fwater_cell.number_format, "0.00")
        
        # Check alternating zebra background formatting:
        # Row 2 (even) -> White background
        # Row 3 (odd) -> Zebra gray background (F2F4F8)
        cell_r2 = ws_data.cell(row=2, column=1)
        cell_r3 = ws_data.cell(row=3, column=1)
        
        self.assertEqual(cell_r2.fill.start_color.rgb, "00FFFFFF")
        self.assertEqual(cell_r3.fill.start_color.rgb, "00F2F4F8")

if __name__ == "__main__":
    unittest.main()
